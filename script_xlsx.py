from openpyxl import load_workbook

workbook = load_workbook("tmp/50mb.xlsx")
sheet = workbook.active
print(sheet.cell(row=2, column=1).value)
